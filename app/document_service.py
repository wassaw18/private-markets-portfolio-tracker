"""
Document Service for file upload, storage, and text extraction
"""
import os
import hashlib
import mimetypes
from typing import Optional, List, Tuple
from pathlib import Path
import shutil
from datetime import datetime
import uuid

# Text extraction libraries
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Configuration
UPLOAD_DIR = Path("uploads/documents")
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
ALLOWED_EXTENSIONS = {
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
    '.txt', '.csv', '.json', '.xml', '.rtf', '.html', '.htm'
}

ALLOWED_MIME_TYPES = {
    'application/pdf',
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-powerpoint',
    'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    'text/plain',
    'text/csv',
    'application/json',
    'application/xml',
    'text/xml',
    'application/rtf',
    'text/html',
    'text/htm'
}

class DocumentService:
    def __init__(self, upload_dir: str = None):
        self.upload_dir = Path(upload_dir) if upload_dir else UPLOAD_DIR
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    def validate_file(self, filename: str, content: bytes) -> Tuple[bool, str]:
        """Validate file extension, size, and MIME type"""
        # Check file size
        if len(content) > MAX_FILE_SIZE:
            return False, f"File size exceeds maximum allowed size of {MAX_FILE_SIZE // (1024*1024)}MB"
        
        # Check file extension
        file_ext = Path(filename).suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            return False, f"File extension '{file_ext}' is not allowed"
        
        # Detect MIME type
        mime_type, _ = mimetypes.guess_type(filename)
        if mime_type and mime_type not in ALLOWED_MIME_TYPES:
            return False, f"MIME type '{mime_type}' is not allowed"
        
        return True, "File is valid"
    
    def calculate_file_hash(self, content: bytes) -> str:
        """Calculate SHA-256 hash of file content"""
        return hashlib.sha256(content).hexdigest()
    
    def generate_unique_filename(self, original_filename: str) -> str:
        """Generate a unique filename to prevent conflicts"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_id = str(uuid.uuid4())[:8]
        file_ext = Path(original_filename).suffix.lower()
        return f"{timestamp}_{unique_id}{file_ext}"
    
    def save_file(self, content: bytes, filename: str) -> Tuple[str, str]:
        """Save file to storage and return file path and generated filename"""
        # Generate unique filename
        unique_filename = self.generate_unique_filename(filename)
        file_path = self.upload_dir / unique_filename
        
        # Create subdirectories by year/month for organization
        year_month = datetime.now().strftime("%Y/%m")
        organized_dir = self.upload_dir / year_month
        organized_dir.mkdir(parents=True, exist_ok=True)
        
        organized_path = organized_dir / unique_filename
        
        # Write file
        with open(organized_path, 'wb') as f:
            f.write(content)
        
        return str(organized_path), unique_filename
    
    def extract_text_content(self, file_path: str, mime_type: str) -> Optional[str]:
        """Extract text content from various file types"""
        try:
            if mime_type == 'application/pdf' and PDF_AVAILABLE:
                return self._extract_pdf_text(file_path)
            elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' and DOCX_AVAILABLE:
                return self._extract_docx_text(file_path)
            elif mime_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] and EXCEL_AVAILABLE:
                return self._extract_excel_text(file_path)
            elif mime_type == 'text/plain':
                return self._extract_plain_text(file_path)
            elif mime_type == 'text/csv':
                return self._extract_csv_text(file_path)
            else:
                return None
        except Exception as e:
            print(f"Error extracting text from {file_path}: {e}")
            return None
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF files"""
        text_content = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text_content.append(page.extract_text())
        return '\n'.join(text_content)
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extract text from DOCX files"""
        doc = DocxDocument(file_path)
        text_content = []
        for paragraph in doc.paragraphs:
            text_content.append(paragraph.text)
        return '\n'.join(text_content)
    
    def _extract_excel_text(self, file_path: str) -> str:
        """Extract text from Excel files"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text_content.append('\t'.join(row_text))
        
        return '\n'.join(text_content)
    
    def _extract_plain_text(self, file_path: str) -> str:
        """Extract text from plain text files"""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    
    def _extract_csv_text(self, file_path: str) -> str:
        """Extract text from CSV files"""
        import csv
        text_content = []
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                text_content.append('\t'.join(row))
        
        return '\n'.join(text_content)
    
    def delete_file(self, file_path: str) -> bool:
        """Delete a file from storage"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return True
            return False
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")
            return False
    
    def get_file_info(self, file_path: str) -> dict:
        """Get file information"""
        path = Path(file_path)
        if not path.exists():
            return {}
        
        stat = path.stat()
        return {
            'size': stat.st_size,
            'modified_time': datetime.fromtimestamp(stat.st_mtime),
            'created_time': datetime.fromtimestamp(stat.st_ctime),
            'mime_type': mimetypes.guess_type(str(path))[0]
        }
    
    def process_upload(self, filename: str, content: bytes, uploaded_by: str = None) -> dict:
        """Process a complete file upload"""
        # Validate file
        is_valid, message = self.validate_file(filename, content)
        if not is_valid:
            raise ValueError(message)
        
        # Calculate hash for deduplication
        file_hash = self.calculate_file_hash(content)
        
        # Detect MIME type
        mime_type, _ = mimetypes.guess_type(filename)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        # Save file
        file_path, unique_filename = self.save_file(content, filename)
        
        # Extract text content
        searchable_content = self.extract_text_content(file_path, mime_type)
        
        return {
            'filename': unique_filename,
            'original_filename': filename,
            'file_path': file_path,
            'file_size': len(content),
            'mime_type': mime_type,
            'file_hash': file_hash,
            'searchable_content': searchable_content,
            'uploaded_by': uploaded_by
        }
    
    def check_duplicate(self, file_hash: str, db_check_function) -> Optional[dict]:
        """Check if a file with the same hash already exists"""
        return db_check_function(file_hash)
    
    def get_storage_stats(self) -> dict:
        """Get storage usage statistics"""
        total_size = 0
        total_files = 0
        
        for root, dirs, files in os.walk(self.upload_dir):
            for file in files:
                file_path = Path(root) / file
                if file_path.exists():
                    total_size += file_path.stat().st_size
                    total_files += 1
        
        return {
            'total_files': total_files,
            'total_size_bytes': total_size,
            'total_size_mb': total_size / (1024 * 1024),
            'upload_directory': str(self.upload_dir)
        }

# Global document service instance
document_service = DocumentService()

def get_document_service() -> DocumentService:
    """Get the global document service instance"""
    return document_service