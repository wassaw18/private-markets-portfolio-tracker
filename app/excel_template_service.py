"""
Excel Template Generation and Processing Service
Professional-grade templates for institutional users
"""
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session
from datetime import datetime, date
import logging

from app import models, schemas, crud
from app.models import CashFlowType

logger = logging.getLogger(__name__)

class ExcelTemplateService:
    """Service for generating professional Excel templates and processing uploads"""
    
    def __init__(self):
        # Professional color scheme
        self.colors = {
            'header': 'FF2E86C1',  # Professional blue
            'secondary': 'FFE8F4F8',  # Light blue
            'success': 'FF27AE60',  # Green
            'warning': 'FFF39C12',  # Orange
            'error': 'FFE74C3C',   # Red
            'border': 'FF5D6D7E'   # Gray
        }
        
        # Standard fonts
        self.fonts = {
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Calibri', size=11, bold=True, color='FF2C3E50'),
            'body': Font(name='Calibri', size=10),
            'instruction': Font(name='Calibri', size=9, italic=True, color='FF7F8C8D')
        }

    def _create_professional_style(self, workbook: Workbook) -> Dict:
        """Create professional styling elements"""
        return {
            'header_fill': PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type="solid"),
            'secondary_fill': PatternFill(start_color=self.colors['secondary'], end_color=self.colors['secondary'], fill_type="solid"),
            'thin_border': Border(
                left=Side(border_style="thin", color=self.colors['border']),
                right=Side(border_style="thin", color=self.colors['border']),
                top=Side(border_style="thin", color=self.colors['border']),
                bottom=Side(border_style="thin", color=self.colors['border'])
            ),
            'center_alignment': Alignment(horizontal="center", vertical="center")
        }

    def generate_nav_template(self, db: Session) -> BytesIO:
        """Generate professional NAV upload template"""
        workbook = Workbook()
        styles = self._create_professional_style(workbook)
        
        # Remove default sheet and create NAV Data sheet
        workbook.remove(workbook.active)
        data_sheet = workbook.create_sheet("NAV Data")
        instructions_sheet = workbook.create_sheet("Instructions")
        validation_sheet = workbook.create_sheet("Validation Data")
        
        # Get investment names for dropdown validation
        investments = crud.get_investments(db, skip=0, limit=1000)
        investment_names = [inv.name for inv in investments]
        print(f"NAV Template: Found {len(investment_names)} investments: {investment_names}")
        
        self._create_nav_data_sheet(data_sheet, investment_names, styles)
        self._create_nav_instructions_sheet(instructions_sheet, styles)
        self._create_validation_data_sheet(validation_sheet, investment_names, styles)
        
        # Set active sheet to data entry
        workbook.active = data_sheet
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def generate_investment_template(self, db: Session) -> BytesIO:
        """Generate professional Investment bulk upload template"""
        workbook = Workbook()
        styles = self._create_professional_style(workbook)
        
        # Remove default sheet and create sheets
        workbook.remove(workbook.active)
        data_sheet = workbook.create_sheet("Investment Data")
        instructions_sheet = workbook.create_sheet("Instructions")
        validation_sheet = workbook.create_sheet("Validation Data")
        
        # Get entities for dropdown validation
        entities = crud.get_entities(db, skip=0, limit=1000)
        entity_names = [f"{entity.name} ({entity.entity_type})" for entity in entities]
        entity_ids = [entity.id for entity in entities]
        print(f"Investment Template: Found {len(entity_names)} entities: {entity_names}")
        
        # Define dropdowns for validation
        asset_classes = ["PRIVATE_EQUITY", "PRIVATE_CREDIT", "REAL_ESTATE", "INFRASTRUCTURE", "HEDGE_FUNDS", "VENTURE_CAPITAL"]
        investment_structures = ["LIMITED_PARTNERSHIP", "FUND_OF_FUNDS", "DIRECT_INVESTMENT", "CO_INVESTMENT", "SEPARATE_ACCOUNT"]
        liquidity_profiles = ["ILLIQUID", "SEMI_LIQUID", "LIQUID"]
        reporting_frequencies = ["MONTHLY", "QUARTERLY", "SEMI_ANNUALLY", "ANNUALLY"]
        risk_ratings = ["LOW", "MEDIUM", "HIGH", "VERY_HIGH"]
        currencies = ["USD", "EUR", "GBP", "JPY"]
        
        self._create_investment_data_sheet(data_sheet, entity_names, styles)
        self._create_investment_instructions_sheet(instructions_sheet, styles)
        self._create_investment_validation_data_sheet(validation_sheet, entity_names, asset_classes, investment_structures, liquidity_profiles, reporting_frequencies, risk_ratings, currencies, styles)
        
        # Set active sheet to data entry
        workbook.active = data_sheet
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def _create_nav_data_sheet(self, sheet, investment_names: List[str], styles: Dict):
        """Create the NAV data entry sheet with validation"""
        # Headers
        headers = [
            "Investment Name", "NAV Date", "NAV Value", "Notes"
        ]
        
        # Header descriptions
        descriptions = [
            "Select from dropdown", "Format: YYYY-MM-DD", "Numeric value > 0", "Optional comments"
        ]
        
        # Apply header styling
        for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = styles['header_fill']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_alignment']
            
            # Add description in row 2
            desc_cell = sheet.cell(row=2, column=col)
            desc_cell.value = desc
            desc_cell.font = self.fonts['instruction']
            desc_cell.fill = styles['secondary_fill']
            desc_cell.border = styles['thin_border']
            desc_cell.alignment = styles['center_alignment']
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25  # Investment Name
        sheet.column_dimensions['B'].width = 15  # NAV Date
        sheet.column_dimensions['C'].width = 15  # NAV Value
        sheet.column_dimensions['D'].width = 30  # Notes
        
        # Add data validation for Investment Name (dropdown)
        if investment_names:
            print(f"Creating dropdown validation for {len(investment_names)} investments")
            investment_validation = DataValidation(
                type="list",
                formula1=f"'Validation Data'!$A$2:$A${len(investment_names)+1}",
                showDropDown=True
            )
            investment_validation.error = "Please select a valid investment name"
            investment_validation.errorTitle = "Invalid Investment"
            sheet.add_data_validation(investment_validation)
            investment_validation.add(f"A3:A1000")
        else:
            print("WARNING: No investments found - dropdown will be empty")
        
        # Add date validation
        date_validation = DataValidation(
            type="date",
            operator="greaterThan",
            formula1="1900-01-01",
            showDropDown=False
        )
        date_validation.error = "Please enter a valid date (YYYY-MM-DD)"
        date_validation.errorTitle = "Invalid Date"
        sheet.add_data_validation(date_validation)
        date_validation.add("B3:B1000")
        
        # Add numeric validation for NAV Value
        value_validation = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            showDropDown=False
        )
        value_validation.error = "NAV Value must be greater than 0"
        value_validation.errorTitle = "Invalid NAV Value"
        sheet.add_data_validation(value_validation)
        value_validation.add("C3:C1000")
        
        # Add sample data
        sample_data = [
            ["Example Fund LP", "2024-12-31", 1000000, "Q4 2024 valuation"],
            ["", "", "", ""],  # Empty row for user input
        ]
        
        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = self.fonts['body']
                cell.border = styles['thin_border']
                if row_idx == 3:  # Sample row - make it stand out
                    cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type="solid")

    def _create_nav_instructions_sheet(self, sheet, styles: Dict):
        """Create comprehensive instructions for NAV upload"""
        instructions = [
            ("NAV Upload Template Instructions", self.fonts['header']),
            ("", None),
            ("Overview:", self.fonts['subheader']),
            ("This template allows you to upload Net Asset Values (NAVs) for your investments.", self.fonts['body']),
            ("Please follow the format exactly to ensure successful import.", self.fonts['body']),
            ("", None),
            ("Column Descriptions:", self.fonts['subheader']),
            ("• Investment Name: Select from the dropdown list of existing investments", self.fonts['body']),
            ("• NAV Date: Enter date in YYYY-MM-DD format (e.g., 2024-12-31)", self.fonts['body']),
            ("• NAV Value: Enter numeric value greater than 0 (e.g., 1000000)", self.fonts['body']),
            ("• Notes: Optional field for additional information", self.fonts['body']),
            ("", None),
            ("Important Notes:", self.fonts['subheader']),
            ("• The investment must already exist in the system", self.fonts['body']),
            ("• NAV Date must be a valid date", self.fonts['body']),
            ("• NAV Value must be positive", self.fonts['body']),
            ("• Duplicate NAV dates for the same investment will be rejected", self.fonts['body']),
            ("• Maximum 1000 records per upload", self.fonts['body']),
            ("", None),
            ("Steps to Use:", self.fonts['subheader']),
            ("1. Switch to the 'NAV Data' tab", self.fonts['body']),
            ("2. Fill in your data starting from row 3", self.fonts['body']),
            ("3. Use the dropdown for Investment Name", self.fonts['body']),
            ("4. Save the file and upload through the web interface", self.fonts['body']),
            ("", None),
            ("Support:", self.fonts['subheader']),
            ("Contact your administrator for assistance or questions", self.fonts['body']),
        ]
        
        for row, (text, font) in enumerate(instructions, 1):
            cell = sheet.cell(row=row, column=1)
            cell.value = text
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        sheet.column_dimensions['A'].width = 80

    def generate_cashflow_template(self, db: Session) -> BytesIO:
        """Generate professional Cash Flow upload template"""
        workbook = Workbook()
        styles = self._create_professional_style(workbook)
        
        # Remove default sheet and create sheets
        workbook.remove(workbook.active)
        data_sheet = workbook.create_sheet("Cash Flow Data")
        instructions_sheet = workbook.create_sheet("Instructions")
        validation_sheet = workbook.create_sheet("Validation Data")
        
        # Get investment names for dropdown validation
        investments = crud.get_investments(db, skip=0, limit=1000)
        investment_names = [inv.name for inv in investments]
        print(f"Cash Flow Template: Found {len(investment_names)} investments: {investment_names}")
        
        # Cash flow types - using the enhanced categories
        cashflow_types = [
            "CAPITAL_CALL",
            "FEES", 
            "YIELD",
            "RETURN_OF_PRINCIPAL",
            "DISTRIBUTION"
        ]
        
        self._create_cashflow_data_sheet(data_sheet, investment_names, cashflow_types, styles)
        self._create_cashflow_instructions_sheet(instructions_sheet, styles)
        self._create_cashflow_validation_data_sheet(validation_sheet, investment_names, cashflow_types, styles)
        
        workbook.active = data_sheet
        
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def _create_cashflow_data_sheet(self, sheet, investment_names: List[str], cashflow_types: List[str], styles: Dict):
        """Create the Cash Flow data entry sheet"""
        headers = [
            "Investment Name", "Date", "Cash Flow Type", "Amount", "Description", "Notes"
        ]
        
        descriptions = [
            "Select from dropdown", 
            "Format: YYYY-MM-DD", 
            "Select type from dropdown",
            "Positive=inflow, Negative=outflow", 
            "Brief description",
            "Optional comments"
        ]
        
        # Apply header styling
        for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = styles['header_fill']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_alignment']
            
            desc_cell = sheet.cell(row=2, column=col)
            desc_cell.value = desc
            desc_cell.font = self.fonts['instruction']
            desc_cell.fill = styles['secondary_fill']
            desc_cell.border = styles['thin_border']
            desc_cell.alignment = styles['center_alignment']
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25  # Investment Name
        sheet.column_dimensions['B'].width = 15  # Date
        sheet.column_dimensions['C'].width = 20  # Cash Flow Type
        sheet.column_dimensions['D'].width = 15  # Amount
        sheet.column_dimensions['E'].width = 25  # Description
        sheet.column_dimensions['F'].width = 30  # Notes
        
        # Investment name validation
        if investment_names:
            print(f"Cash Flow: Creating investment dropdown with {len(investment_names)} options")
            investment_validation = DataValidation(
                type="list",
                formula1=f"'Validation Data'!$A$2:$A${len(investment_names)+1}",
                showDropDown=True
            )
            sheet.add_data_validation(investment_validation)
            investment_validation.add("A3:A1000")
        else:
            print("WARNING: Cash Flow - No investments found for dropdown")
        
        # Date validation
        date_validation = DataValidation(
            type="date",
            operator="greaterThan",
            formula1="1900-01-01",
            showDropDown=False
        )
        sheet.add_data_validation(date_validation)
        date_validation.add("B3:B1000")
        
        # Cash flow type validation
        print(f"Cash Flow: Creating cash flow type dropdown with {len(cashflow_types)} options: {cashflow_types}")
        cashflow_validation = DataValidation(
            type="list",
            formula1=f"'Validation Data'!$B$2:$B${len(cashflow_types)+1}",
            showDropDown=True
        )
        sheet.add_data_validation(cashflow_validation)
        cashflow_validation.add("C3:C1000")
        
        # Amount validation (numeric)
        amount_validation = DataValidation(
            type="decimal",
            operator="notEqual",
            formula1=0,
            showDropDown=False
        )
        amount_validation.error = "Amount cannot be zero"
        amount_validation.errorTitle = "Invalid Amount"
        sheet.add_data_validation(amount_validation)
        amount_validation.add("D3:D1000")
        
        # Add sample data
        sample_data = [
            ["Example Fund LP", "2024-12-31", "CAPITAL_CALL", -500000, "Q4 2024 capital call", ""],
            ["Example Fund LP", "2024-12-31", "DISTRIBUTION", 750000, "Q4 2024 distribution", "Income + capital"],
            ["", "", "", "", "", ""],  # Empty row
        ]
        
        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = self.fonts['body']
                cell.border = styles['thin_border']
                if row_idx <= 4:  # Sample rows
                    cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type="solid")

    def _create_cashflow_instructions_sheet(self, sheet, styles: Dict):
        """Create comprehensive instructions for Cash Flow upload"""
        instructions = [
            ("Cash Flow Upload Template Instructions", self.fonts['header']),
            ("", None),
            ("Overview:", self.fonts['subheader']),
            ("This template allows you to upload cash flow transactions for your investments.", self.fonts['body']),
            ("Each row represents a single cash flow transaction.", self.fonts['body']),
            ("", None),
            ("Column Descriptions:", self.fonts['subheader']),
            ("• Investment Name: Select from dropdown of existing investments", self.fonts['body']),
            ("• Date: Transaction date in YYYY-MM-DD format", self.fonts['body']),
            ("• Cash Flow Type: Select from dropdown:", self.fonts['body']),
            ("  - CAPITAL_CALL: Money called from investors", self.fonts['body']),
            ("  - FEES: Management fees and expenses", self.fonts['body']),
            ("  - YIELD: Income distributions (dividends, interest)", self.fonts['body']),
            ("  - RETURN_OF_PRINCIPAL: Return of original capital", self.fonts['body']),
            ("  - DISTRIBUTION: General distributions (mixed types)", self.fonts['body']),
            ("• Amount: Positive for money received, negative for money paid", self.fonts['body']),
            ("• Description: Brief description of the transaction", self.fonts['body']),
            ("• Notes: Additional comments or details", self.fonts['body']),
            ("", None),
            ("Amount Convention:", self.fonts['subheader']),
            ("• Positive amounts: Money flowing TO you (distributions, returns)", self.fonts['body']),
            ("• Negative amounts: Money flowing FROM you (capital calls, fees)", self.fonts['body']),
            ("", None),
            ("Examples:", self.fonts['subheader']),
            ("• Capital Call: Amount = -500,000 (you pay money)", self.fonts['body']),
            ("• Distribution: Amount = 750,000 (you receive money)", self.fonts['body']),
            ("• Management Fee: Amount = -25,000 (you pay fee)", self.fonts['body']),
            ("", None),
            ("Data Requirements:", self.fonts['subheader']),
            ("• Investment must exist in the system", self.fonts['body']),
            ("• Date must be valid", self.fonts['body']),
            ("• Cash Flow Type must be selected from dropdown", self.fonts['body']),
            ("• Amount cannot be zero", self.fonts['body']),
            ("• Description is recommended for clarity", self.fonts['body']),
            ("• Maximum 1000 records per upload", self.fonts['body']),
            ("", None),
            ("Processing Notes:", self.fonts['subheader']),
            ("• Investment summaries will be automatically updated", self.fonts['body']),
            ("• All transactions are processed as a batch", self.fonts['body']),
            ("• If any error occurs, entire batch is rejected", self.fonts['body']),
            ("• Detailed error reporting will identify issues", self.fonts['body']),
            ("", None),
            ("Steps to Use:", self.fonts['subheader']),
            ("1. Switch to 'Cash Flow Data' tab", self.fonts['body']),
            ("2. Fill in your data starting from row 3", self.fonts['body']),
            ("3. Use dropdowns for Investment Name and Cash Flow Type", self.fonts['body']),
            ("4. Follow amount conventions (positive/negative)", self.fonts['body']),
            ("5. Save file and upload through web interface", self.fonts['body']),
        ]
        
        for row, (text, font) in enumerate(instructions, 1):
            cell = sheet.cell(row=row, column=1)
            cell.value = text
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        sheet.column_dimensions['A'].width = 80

    def _create_validation_data_sheet(self, sheet, investment_names: List[str], styles: Dict):
        """Create validation data sheet for dropdowns"""
        # Investment names in column A
        sheet.cell(row=1, column=1, value="Investment Names")
        sheet.cell(row=1, column=1).font = self.fonts['subheader']
        
        for idx, name in enumerate(investment_names, 2):
            sheet.cell(row=idx, column=1, value=name)
        
        sheet.column_dimensions['A'].width = 30

    def _create_cashflow_validation_data_sheet(self, sheet, investment_names: List[str], cashflow_types: List[str], styles: Dict):
        """Create validation data for Cash Flow template"""
        print(f"Creating validation sheet with {len(investment_names)} investments and {len(cashflow_types)} types")
        
        # Investment names in column A
        sheet.cell(row=1, column=1, value="Investment Names")
        sheet.cell(row=1, column=1).font = self.fonts['subheader']
        
        for idx, name in enumerate(investment_names, 2):
            sheet.cell(row=idx, column=1, value=name)
            print(f"Added investment to validation: {name}")
        
        # Cash flow types in column B
        sheet.cell(row=1, column=2, value="Cash Flow Types")
        sheet.cell(row=1, column=2).font = self.fonts['subheader']
        
        for idx, cf_type in enumerate(cashflow_types, 2):
            sheet.cell(row=idx, column=2, value=cf_type)
            print(f"Added cash flow type to validation: {cf_type}")
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20

    def _create_investment_data_sheet(self, sheet, entity_names: List[str], styles: Dict):
        """Create the Investment data entry sheet with validation"""
        # Headers - Required fields marked with *
        headers = [
            "Investment Name*", "Asset Class*", "Investment Structure*", "Entity*", "Strategy*", 
            "Vintage Year*", "Commitment Amount*", "Commitment Date*", "Called Amount", "Currency",
            "Management Fee (%)", "Performance Fee (%)", "Hurdle Rate (%)", "Preferred Return (%)",
            "Contact Person", "Email", "Phone", "Address", "Fund Size", "Target Return (%)",
            "Investment Period (years)", "Fund Life (years)", "Reporting Frequency", "Geography",
            "Sector Focus", "Risk Rating", "ESG Focus", "Other Fees", "Key Terms", "Due Diligence Notes"
        ]
        
        # Descriptions
        descriptions = [
            "Required", "Required", "Required", "Required", "Required",
            "Required", "Required", "Required", "Optional", "Optional",
            "Optional", "Optional", "Optional", "Optional",
            "Optional", "Optional", "Optional", "Optional", "Optional", "Optional",
            "Optional", "Optional", "Optional", "Optional",
            "Optional", "Optional", "Optional", "Optional", "Optional", "Optional"
        ]
        
        # Apply header styling
        for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = styles['header_fill']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_alignment']
            
            # Add description in row 2
            desc_cell = sheet.cell(row=2, column=col)
            desc_cell.value = desc
            desc_cell.font = self.fonts['instruction']
            desc_cell.fill = styles['secondary_fill']
            desc_cell.border = styles['thin_border']
            desc_cell.alignment = styles['center_alignment']
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 18
        
        # Wider columns for text fields
        sheet.column_dimensions['A'].width = 25  # Investment Name
        sheet.column_dimensions['D'].width = 25  # Entity
        sheet.column_dimensions['E'].width = 20  # Strategy
        
        # Add sample data
        sample_data = [
            "Example Fund LP", "PRIVATE_EQUITY", "LIMITED_PARTNERSHIP", entity_names[0] if entity_names else "Create Entity First",
            "Growth Buyout", "2024", "5000000", "2024-01-15", "1000000", "USD",
            "2.0", "20.0", "8.0", "8.0",
            "John Smith", "john@fund.com", "+1-555-123-4567", "123 Main St", "500000000", "15.0",
            "5", "10", "QUARTERLY", "North America",
            "Technology", "MEDIUM", "ESG Focus", "25000", "Key terms here", "DD notes here"
        ]
        
        for col_idx, value in enumerate(sample_data, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = value
            cell.font = self.fonts['body']
            cell.border = styles['thin_border']
            cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type="solid")

    def _create_investment_instructions_sheet(self, sheet, styles: Dict):
        """Create comprehensive instructions for Investment upload"""
        instructions = [
            ("Investment Bulk Upload Template Instructions", self.fonts['header']),
            ("", None),
            ("Overview:", self.fonts['subheader']),
            ("This template allows you to upload multiple investments at once.", self.fonts['body']),
            ("Required fields are marked with * and must be completed.", self.fonts['body']),
            ("", None),
            ("Required Fields (*) - Must be completed:", self.fonts['subheader']),
            ("• Investment Name: Unique name for the investment", self.fonts['body']),
            ("• Asset Class: Select from dropdown", self.fonts['body']),
            ("• Investment Structure: Select from dropdown", self.fonts['body']),
            ("• Entity: Select from existing entities", self.fonts['body']),
            ("• Strategy: Investment strategy description", self.fonts['body']),
            ("• Vintage Year: Year of investment (YYYY)", self.fonts['body']),
            ("• Commitment Amount: Total committed capital", self.fonts['body']),
            ("• Commitment Date: Date in YYYY-MM-DD format", self.fonts['body']),
            ("", None),
            ("Optional Fields - Can be updated later:", self.fonts['subheader']),
            ("• Financial: Management fees, performance fees, hurdle rates", self.fonts['body']),
            ("• Contact: Person, email, phone, address information", self.fonts['body']),
            ("• Operational: Fund size, target return, investment periods", self.fonts['body']),
            ("• Legal: Geography, sector focus, risk rating, ESG focus", self.fonts['body']),
            ("• Administrative: Other fees, key terms, due diligence notes", self.fonts['body']),
            ("", None),
            ("Data Format Guidelines:", self.fonts['subheader']),
            ("• Dates: Use YYYY-MM-DD format (e.g., 2024-12-31)", self.fonts['body']),
            ("• Percentages: Enter as numbers (e.g., 2.5 for 2.5%)", self.fonts['body']),
            ("• Amounts: Enter numeric values without commas", self.fonts['body']),
            ("• Years: Enter as whole numbers", self.fonts['body']),
            ("• Text Fields: Keep descriptions concise", self.fonts['body']),
            ("", None),
            ("Steps to Use:", self.fonts['subheader']),
            ("1. Switch to the 'Investment Data' tab", self.fonts['body']),
            ("2. Fill in required fields (marked with *) for each investment", self.fonts['body']),
            ("3. Use dropdowns for standardized fields", self.fonts['body']),
            ("4. Optional fields can be left blank and updated later", self.fonts['body']),
            ("5. Save the file and upload through the web interface", self.fonts['body']),
            ("", None),
            ("Important Notes:", self.fonts['subheader']),
            ("• Entity must exist before creating investments", self.fonts['body']),
            ("• Investment names must be unique", self.fonts['body']),
            ("• Called Amount cannot exceed Commitment Amount", self.fonts['body']),
            ("• Maximum 1000 investments per upload", self.fonts['body']),
        ]
        
        for row, (text, font) in enumerate(instructions, 1):
            cell = sheet.cell(row=row, column=1)
            cell.value = text
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        sheet.column_dimensions['A'].width = 80

    def _create_investment_validation_data_sheet(self, sheet, entity_names: List[str], asset_classes: List[str], investment_structures: List[str], liquidity_profiles: List[str], reporting_frequencies: List[str], risk_ratings: List[str], currencies: List[str], styles: Dict):
        """Create validation data for Investment template dropdowns"""
        
        # Column headers
        headers = ["Entities", "Asset Classes", "Investment Structures", "Liquidity Profiles", "Reporting Frequencies", "Risk Ratings", "Currencies"]
        data_lists = [entity_names, asset_classes, investment_structures, liquidity_profiles, reporting_frequencies, risk_ratings, currencies]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.fonts['subheader']
        
        # Add data for each column
        for col, data_list in enumerate(data_lists, 1):
            for idx, item in enumerate(data_list, 2):
                sheet.cell(row=idx, column=col, value=item)
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 25

    def _create_nav_data_sheet(self, sheet, investment_names: List[str], styles: Dict):
        """Create the NAV data entry sheet with validation"""
        # Headers
        headers = [
            "Investment Name", "NAV Date", "NAV Value", "Notes"
        ]
        
        # Header descriptions
        descriptions = [
            "Select from dropdown", "Format: YYYY-MM-DD", "Numeric value > 0", "Optional comments"
        ]
        
        # Apply header styling
        for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = styles['header_fill']
            cell.border = styles['thin_border']
            cell.alignment = styles['center_alignment']
            
            # Add description in row 2
            desc_cell = sheet.cell(row=2, column=col)
            desc_cell.value = desc
            desc_cell.font = self.fonts['instruction']
            desc_cell.fill = styles['secondary_fill']
            desc_cell.border = styles['thin_border']
            desc_cell.alignment = styles['center_alignment']
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25  # Investment Name
        sheet.column_dimensions['B'].width = 15  # NAV Date
        sheet.column_dimensions['C'].width = 15  # NAV Value
        sheet.column_dimensions['D'].width = 30  # Notes
        
        # Add data validation for Investment Name (dropdown)
        if investment_names:
            print(f"Creating dropdown validation for {len(investment_names)} investments")
            investment_validation = DataValidation(
                type="list",
                formula1=f"'Validation Data'!$A$2:$A${len(investment_names)+1}",
                showDropDown=True
            )
            investment_validation.error = "Please select a valid investment name"
            investment_validation.errorTitle = "Invalid Investment"
            sheet.add_data_validation(investment_validation)
            investment_validation.add(f"A3:A1000")
        else:
            print("WARNING: No investments found - dropdown will be empty")
        
        # Add date validation
        date_validation = DataValidation(
            type="date",
            operator="greaterThan",
            formula1="1900-01-01",
            showDropDown=False
        )
        date_validation.error = "Please enter a valid date (YYYY-MM-DD)"
        date_validation.errorTitle = "Invalid Date"
        sheet.add_data_validation(date_validation)
        date_validation.add("B3:B1000")
        
        # Add numeric validation for NAV Value
        value_validation = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1=0,
            showDropDown=False
        )
        value_validation.error = "NAV Value must be greater than 0"
        value_validation.errorTitle = "Invalid NAV Value"
        sheet.add_data_validation(value_validation)
        value_validation.add("C3:C1000")
        
        # Add sample data
        sample_data = [
            ["Example Fund LP", "2024-12-31", 1000000, "Q4 2024 valuation"],
            ["", "", "", ""],  # Empty row for user input
        ]
        
        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = self.fonts['body']
                cell.border = styles['thin_border']
                if row_idx == 3:  # Sample row - make it stand out
                    cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type="solid")

class BulkUploadProcessor:
    """Process bulk uploads from Excel templates"""
    
    @staticmethod
    def process_nav_upload(file_content: bytes, filename: str, db: Session) -> BulkUploadResult:
        """Process NAV bulk upload"""
        result = BulkUploadResult()
        
        try:
            # Read Excel file
            df = pd.read_excel(BytesIO(file_content), sheet_name='NAV Data', skiprows=1)
            
            logger.info(f"Processing {len(df)} NAV records from {filename}")
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            # Get investment name mapping
            investments = crud.get_investments(db, skip=0, limit=1000)
            investment_map = {inv.name: inv.id for inv in investments}
            
            for index, row in df.iterrows():
                row_num = index + 3  # Account for header rows
                
                # Skip empty rows
                if pd.isna(row.get('Investment Name')):
                    continue
                
                try:
                    # Validate investment exists
                    investment_name = str(row['Investment Name']).strip()
                    if investment_name not in investment_map:
                        result.add_error(row_num, f"Investment '{investment_name}' not found")
                        continue
                    
                    investment_id = investment_map[investment_name]
                    
                    # Validate NAV date
                    nav_date = row['NAV Date']
                    if pd.isna(nav_date):
                        result.add_error(row_num, "NAV Date is required")
                        continue
                    
                    if isinstance(nav_date, str):
                        try:
                            nav_date = datetime.strptime(nav_date, '%Y-%m-%d').date()
                        except ValueError:
                            result.add_error(row_num, "Invalid NAV Date format. Use YYYY-MM-DD")
                            continue
                    elif isinstance(nav_date, datetime):
                        nav_date = nav_date.date()
                    
                    # Validate NAV value
                    nav_value = row['NAV Value']
                    if pd.isna(nav_value) or nav_value <= 0:
                        result.add_error(row_num, "NAV Value must be greater than 0")
                        continue
                    
                    # Check for duplicate
                    existing_valuation = db.query(models.Valuation).filter(
                        models.Valuation.investment_id == investment_id,
                        models.Valuation.date == nav_date
                    ).first()
                    
                    if existing_valuation:
                        result.add_warning(row_num, f"NAV for {nav_date} already exists, updating value")
                        existing_valuation.value = float(nav_value)
                        existing_valuation.notes = str(row.get('Notes', ''))
                        db.commit()
                    else:
                        # Create new valuation
                        valuation = models.Valuation(
                            investment_id=investment_id,
                            date=nav_date,
                            value=float(nav_value),
                            notes=str(row.get('Notes', '')) if not pd.isna(row.get('Notes')) else ""
                        )
                        db.add(valuation)
                        db.commit()
                    
                    result.add_success()
                    
                except Exception as e:
                    result.add_error(row_num, f"Processing error: {str(e)}")
            
            result.message = f"Processed {result.success_count} NAV records successfully"
            if result.error_count > 0:
                result.message += f" with {result.error_count} errors"
                
        except Exception as e:
            result.add_error(0, f"File processing error: {str(e)}")
            
        return result

    @staticmethod
    def process_cashflow_upload(file_content: bytes, filename: str, db: Session) -> BulkUploadResult:
        """Process Cash Flow bulk upload"""
        result = BulkUploadResult()
        
        try:
            # Read Excel file
            df = pd.read_excel(BytesIO(file_content), sheet_name='Cash Flow Data', skiprows=1)
            
            logger.info(f"Processing {len(df)} cash flow records from {filename}")
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            # Get investment name mapping
            investments = crud.get_investments(db, skip=0, limit=1000)
            investment_map = {inv.name: inv.id for inv in investments}
            
            # Cash flow type mapping
            cf_type_map = {
                'CAPITAL_CALL': CashFlowType.CAPITAL_CALL,
                'FEES': CashFlowType.FEES,
                'YIELD': CashFlowType.YIELD,
                'RETURN_OF_PRINCIPAL': CashFlowType.RETURN_OF_PRINCIPAL,
                'DISTRIBUTION': CashFlowType.DISTRIBUTION
            }
            
            for index, row in df.iterrows():
                row_num = index + 3
                
                # Skip empty rows
                if pd.isna(row.get('Investment Name')):
                    continue
                
                try:
                    # Validate investment exists
                    investment_name = str(row['Investment Name']).strip()
                    if investment_name not in investment_map:
                        result.add_error(row_num, f"Investment '{investment_name}' not found")
                        continue
                    
                    investment_id = investment_map[investment_name]
                    
                    # Validate date
                    cf_date = row['Date']
                    if pd.isna(cf_date):
                        result.add_error(row_num, "Date is required")
                        continue
                    
                    if isinstance(cf_date, str):
                        try:
                            cf_date = datetime.strptime(cf_date, '%Y-%m-%d').date()
                        except ValueError:
                            result.add_error(row_num, "Invalid Date format. Use YYYY-MM-DD")
                            continue
                    elif isinstance(cf_date, datetime):
                        cf_date = cf_date.date()
                    
                    # Validate cash flow type
                    cf_type_str = str(row['Cash Flow Type']).strip().upper()
                    if cf_type_str not in cf_type_map:
                        result.add_error(row_num, f"Invalid Cash Flow Type: {cf_type_str}")
                        continue
                    
                    cf_type = cf_type_map[cf_type_str]
                    
                    # Validate amount
                    amount = row['Amount']
                    if pd.isna(amount) or amount == 0:
                        result.add_error(row_num, "Amount cannot be zero or empty")
                        continue
                    
                    # Create cash flow
                    cashflow = models.CashFlow(
                        investment_id=investment_id,
                        date=cf_date,
                        amount=float(amount),
                        type=cf_type,
                        notes=str(row.get('Notes', '')) if not pd.isna(row.get('Notes')) else ""
                    )
                    
                    db.add(cashflow)
                    result.add_success()
                    
                except Exception as e:
                    result.add_error(row_num, f"Processing error: {str(e)}")
            
            # Commit all changes if no errors
            if result.error_count == 0:
                db.commit()
                result.message = f"Successfully processed {result.success_count} cash flow records"
                
                # Update investment summaries
                try:
                    BulkUploadProcessor._update_investment_summaries(db, list(investment_map.values()))
                except Exception as e:
                    result.add_warning(0, f"Investment summary update failed: {str(e)}")
            else:
                db.rollback()
                result.message = f"Upload failed due to {result.error_count} errors. No records were saved."
                
        except Exception as e:
            result.add_error(0, f"File processing error: {str(e)}")
            
        return result

    @staticmethod
    def _update_investment_summaries(db: Session, investment_ids: List[int]):
        """Update investment summaries after cash flow changes"""
        for investment_id in set(investment_ids):  # Remove duplicates
            investment = crud.get_investment(db, investment_id)
            if investment:
                # Recalculate called amount and fees
                cashflows = crud.get_investment_cashflows(db, investment_id)
                
                called_amount = sum(
                    abs(cf.amount) for cf in cashflows 
                    if cf.type == CashFlowType.CAPITAL_CALL and cf.amount < 0
                )
                
                fees = sum(
                    abs(cf.amount) for cf in cashflows 
                    if cf.type == CashFlowType.FEES and cf.amount < 0
                )
                
                investment.called_amount = called_amount
                investment.fees = fees
        
        db.commit()

# Global service instance
excel_template_service = ExcelTemplateService()