#!/usr/bin/env python3
"""
Create Excel templates for PitchBook benchmark data import.
This script generates structured Excel files with multiple worksheets,
data validation, and clear instructions for data entry.
"""

import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_pitchbook_excel_template():
    """Create comprehensive Excel template for PitchBook data import."""

    # Create workbook and remove default sheet
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    instruction_font = Font(italic=True, color="666666")
    example_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================
    # 1. INSTRUCTIONS WORKSHEET
    # ========================================

    instructions_ws = wb.create_sheet("Instructions")

    instructions_content = [
        ["PitchBook Benchmark Data Import Template", "", ""],
        ["", "", ""],
        ["PURPOSE:", "Import quarterly PitchBook benchmark data into the private markets database", ""],
        ["", "", ""],
        ["WORKSHEETS:", "", ""],
        ["• Instructions", "This sheet - read first", ""],
        ["• Performance Data", "Vintage year performance metrics (IRR, PME, TVPI, DPI, RVPI)", ""],
        ["• Quarterly Returns", "Quarterly return data by asset class", ""],
        ["• Reference Data", "Valid values for dropdowns and validation", ""],
        ["", "", ""],
        ["BEFORE YOU START:", "", ""],
        ["1. Read all instructions carefully", "", ""],
        ["2. Review the Reference Data sheet for valid values", "", ""],
        ["3. Use only the predefined asset classes and metric codes", "", ""],
        ["4. Enter data in decimal format (0.1420 = 14.20%)", "", ""],
        ["5. Ensure quartile order: Top ≥ Median ≥ Bottom", "", ""],
        ["", "", ""],
        ["DATA FORMAT EXAMPLES:", "", ""],
        ["IRR of 14.20%", "Enter as: 0.1420", ""],
        ["Multiple of 2.45x", "Enter as: 2.45", ""],
        ["Return of -2.50%", "Enter as: -0.0250", ""],
        ["", "", ""],
        ["IMPORT PROCESS:", "", ""],
        ["1. Fill in Performance Data sheet", "", ""],
        ["2. Fill in Quarterly Returns sheet", "", ""],
        ["3. Save file as .xlsx", "", ""],
        ["4. Upload to import system", "", ""],
        ["5. Review import log for any errors", "", ""]
    ]

    for row_idx, row_data in enumerate(instructions_content, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = instructions_ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if row_idx == 1:  # Title row
                cell.font = Font(bold=True, size=16, color="366092")
            elif row_data[0] and row_data[0].endswith(":"):  # Section headers
                cell.font = Font(bold=True, color="366092")

    # Adjust column widths
    instructions_ws.column_dimensions['A'].width = 25
    instructions_ws.column_dimensions['B'].width = 50
    instructions_ws.column_dimensions['C'].width = 20

    # ========================================
    # 2. REFERENCE DATA WORKSHEET
    # ========================================

    reference_ws = wb.create_sheet("Reference Data")

    # Asset classes
    asset_classes = [
        ["Asset Classes (use exact values)", ""],
        ["private_equity", "Private Equity"],
        ["venture_capital", "Venture Capital"],
        ["real_estate", "Real Estate"],
        ["real_assets", "Real Assets"],
        ["private_debt", "Private Debt"],
        ["fund_of_funds", "Fund of Funds"],
        ["secondaries", "Secondaries"]
    ]

    # Metric codes
    metric_codes = [
        ["Metric Codes (use exact values)", ""],
        ["IRR", "Internal Rate of Return"],
        ["PME", "Public Market Equivalent"],
        ["TVPI", "Total Value to Paid-In"],
        ["DPI", "Distributions to Paid-In"],
        ["RVPI", "Residual Value to Paid-In"]
    ]

    # Write asset classes
    for row_idx, (code, name) in enumerate(asset_classes, 1):
        reference_ws.cell(row=row_idx, column=1, value=code).font = header_font if row_idx == 1 else None
        reference_ws.cell(row=row_idx, column=2, value=name).font = header_font if row_idx == 1 else None
        if row_idx == 1:
            reference_ws.cell(row=row_idx, column=1).fill = header_fill
            reference_ws.cell(row=row_idx, column=2).fill = header_fill

    # Write metric codes
    for row_idx, (code, name) in enumerate(metric_codes, 1):
        reference_ws.cell(row=row_idx, column=4, value=code).font = header_font if row_idx == 1 else None
        reference_ws.cell(row=row_idx, column=5, value=name).font = header_font if row_idx == 1 else None
        if row_idx == 1:
            reference_ws.cell(row=row_idx, column=4).fill = header_fill
            reference_ws.cell(row=row_idx, column=5).fill = header_fill

    # Column headers for examples
    reference_ws.cell(row=10, column=1, value="Example Report Periods").font = header_font
    reference_ws.cell(row=10, column=1).fill = header_fill
    reference_ws.cell(row=11, column=1, value="Q4-2024")
    reference_ws.cell(row=12, column=1, value="Q1-2025")
    reference_ws.cell(row=13, column=1, value="Q2-2025")

    reference_ws.cell(row=10, column=3, value="Example Vintage Years").font = header_font
    reference_ws.cell(row=10, column=3).fill = header_fill
    for year in range(2000, 2025, 5):
        reference_ws.cell(row=11 + (year-2000)//5, column=3, value=year)

    # Adjust column widths
    reference_ws.column_dimensions['A'].width = 20
    reference_ws.column_dimensions['B'].width = 25
    reference_ws.column_dimensions['C'].width = 20
    reference_ws.column_dimensions['D'].width = 20
    reference_ws.column_dimensions['E'].width = 25

    # ========================================
    # 3. PERFORMANCE DATA WORKSHEET
    # ========================================

    perf_ws = wb.create_sheet("Performance Data")

    # Headers
    perf_headers = [
        "report_period", "asset_class", "metric_code", "vintage_year",
        "top_quartile_value", "median_value", "bottom_quartile_value",
        "sample_size", "fund_count", "methodology_notes"
    ]

    # Write headers
    for col_idx, header in enumerate(perf_headers, 1):
        cell = perf_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Add instruction row
    instruction_row = [
        "Q4-2024", "private_equity", "IRR", "2020",
        "0.1850", "0.1420", "0.0980",
        "125", "125", "Net IRR as of Q4 2024"
    ]

    for col_idx, value in enumerate(instruction_row, 1):
        cell = perf_ws.cell(row=2, column=col_idx, value=value)
        cell.fill = example_fill
        cell.border = border

    # Add comment to first data row
    perf_ws.cell(row=2, column=1).comment = openpyxl.comments.Comment(
        "This is an example row. Replace with actual data. Delete this row when done.",
        "Template"
    )

    # Set up data validation for asset classes
    asset_class_validation = DataValidation(
        type="list",
        formula1='"private_equity,venture_capital,real_estate,real_assets,private_debt,fund_of_funds,secondaries"',
        allow_blank=False
    )
    perf_ws.add_data_validation(asset_class_validation)
    asset_class_validation.add("B3:B1000")

    # Set up data validation for metric codes
    metric_validation = DataValidation(
        type="list",
        formula1='"IRR,PME,TVPI,DPI,RVPI"',
        allow_blank=False
    )
    perf_ws.add_data_validation(metric_validation)
    metric_validation.add("C3:C1000")

    # Adjust column widths
    column_widths = [12, 15, 12, 12, 16, 14, 18, 12, 12, 25]
    for col_idx, width in enumerate(column_widths, 1):
        perf_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ========================================
    # 4. QUARTERLY RETURNS WORKSHEET
    # ========================================

    quarterly_ws = wb.create_sheet("Quarterly Returns")

    # Headers
    quarterly_headers = [
        "report_period", "asset_class", "quarter_year", "quarter_date",
        "top_quartile_return", "median_return", "bottom_quartile_return", "sample_size"
    ]

    # Write headers
    for col_idx, header in enumerate(quarterly_headers, 1):
        cell = quarterly_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Add instruction row
    quarterly_instruction_row = [
        "Q4-2024", "private_equity", "Q1-2020", "2020-01-01",
        "0.0650", "0.0450", "0.0250", "245"
    ]

    for col_idx, value in enumerate(quarterly_instruction_row, 1):
        cell = quarterly_ws.cell(row=2, column=col_idx, value=value)
        cell.fill = example_fill
        cell.border = border
        if col_idx == 4:  # quarter_date column
            cell.value = datetime.strptime(value, "%Y-%m-%d").date()

    # Add comment to first data row
    quarterly_ws.cell(row=2, column=1).comment = openpyxl.comments.Comment(
        "This is an example row. Replace with actual data. Delete this row when done.",
        "Template"
    )

    # Set up data validation for asset classes
    quarterly_ws.add_data_validation(asset_class_validation)
    asset_class_validation.add("B3:B1000")

    # Adjust column widths
    quarterly_column_widths = [12, 15, 12, 12, 18, 14, 20, 12]
    for col_idx, width in enumerate(quarterly_column_widths, 1):
        quarterly_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ========================================
    # 5. SAVE THE FILE
    # ========================================

    # Create output directory if it doesn't exist
    output_dir = "/home/will/Tmux-Orchestrator/private-markets-tracker/docs/benchmarks/templates"
    os.makedirs(output_dir, exist_ok=True)

    # Save the file
    output_path = os.path.join(output_dir, "PitchBook_Benchmark_Import_Template.xlsx")
    wb.save(output_path)

    print(f"✅ Excel template created successfully: {output_path}")
    print(f"📊 Worksheets included:")
    print(f"   • Instructions - Usage guidelines and examples")
    print(f"   • Reference Data - Valid values for dropdowns")
    print(f"   • Performance Data - Vintage year metrics template")
    print(f"   • Quarterly Returns - Quarterly return data template")
    print(f"🔄 Template includes data validation and formatting")

    return output_path

if __name__ == "__main__":
    try:
        create_pitchbook_excel_template()
    except ImportError as e:
        print(f"❌ Missing required library: {e}")
        print("💡 Install with: pip install openpyxl pandas")
    except Exception as e:
        print(f"❌ Error creating template: {e}")