#!/usr/bin/env python3
"""
Create sample Excel files for testing Private Markets Portfolio Tracker
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date

def create_sample_holdings():
    """Create sample investment holdings data"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Holdings Data"
    
    # Headers
    headers = [
        'name', 'asset_class', 'investment_structure', 'owner', 'strategy',
        'vintage_year', 'commitment_amount', 'called_amount', 'fees',
        'first_close_date', 'final_close_date', 'fund_term_years',
        'investment_period_years', 'fund_size', 'liquidity_profile',
        'geography', 'sector', 'notes'
    ]
    
    # Sample data
    holdings_data = [
        ['Acme Venture Fund III', 'VENTURE_CAPITAL', 'LIMITED_PARTNERSHIP', 'Smith Family Trust', 'Early Stage Technology',
         2022, 5000000, 2500000, 125000, '2022-03-15', '2022-09-30', 10, 5, 250000000, 'ILLIQUID',
         'North America', 'Technology', 'Top-tier VC fund focused on AI and enterprise software'],
        
        ['Global Private Equity Partners', 'PRIVATE_EQUITY', 'LIMITED_PARTNERSHIP', 'Johnson LLC', 'Mid-Market Buyout',
         2021, 10000000, 7500000, 375000, '2021-06-01', '2021-12-15', 10, 5, 1500000000, 'ILLIQUID',
         'Global', 'Healthcare', 'Established PE firm with strong healthcare expertise'],
        
        ['Prime Real Estate Fund IV', 'REAL_ESTATE', 'LIMITED_PARTNERSHIP', 'Wilson Trust', 'Core-Plus Real Estate',
         2023, 3000000, 1200000, 60000, '2023-01-20', '2023-07-31', 8, 4, 800000000, 'SEMI_LIQUID',
         'United States', 'Real Estate', 'Focus on office and industrial properties in major metros'],
        
        ['Credit Opportunities Fund II', 'PRIVATE_CREDIT', 'LIMITED_PARTNERSHIP', 'Davis Family Office', 'Direct Lending',
         2022, 7500000, 5250000, 262500, '2022-05-10', '2022-11-30', 7, 4, 1200000000, 'SEMI_LIQUID',
         'North America', 'Financial Services', 'Senior secured lending to middle market companies'],
        
        ['Emerging Markets Infrastructure', 'REAL_ASSETS', 'LIMITED_PARTNERSHIP', 'Brown Corporation', 'Infrastructure Development',
         2021, 8000000, 6400000, 320000, '2021-08-15', '2022-02-28', 12, 6, 2000000000, 'ILLIQUID',
         'Asia Pacific', 'Infrastructure', 'Focus on renewable energy and transportation infrastructure']
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row_idx, row_data in enumerate(holdings_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_name].width = adjusted_width
    
    # Save file
    output_path = 'sample_holdings_data.xlsx'
    workbook.save(output_path)
    print(f"✅ Sample holdings file created: {output_path}")
    return output_path

def create_sample_cash_flows():
    """Create sample cash flow data"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cash Flow Data"
    
    # Headers
    headers = ['investment_name', 'date', 'amount', 'type', 'notes']
    
    # Sample data
    cash_flows_data = [
        ['Acme Venture Fund III', '2022-04-15', -1000000, 'CAPITAL_CALL', 'Initial capital call for fund deployment'],
        ['Acme Venture Fund III', '2022-08-20', -750000, 'CAPITAL_CALL', 'Second capital call'],
        ['Acme Venture Fund III', '2023-12-15', 250000, 'DISTRIBUTION', 'Distribution from portfolio company exit'],
        
        ['Global Private Equity Partners', '2021-07-01', -2500000, 'CAPITAL_CALL', 'Initial funding'],
        ['Global Private Equity Partners', '2021-10-15', -2000000, 'CAPITAL_CALL', 'Follow-on capital call'],
        ['Global Private Equity Partners', '2022-06-30', -1500000, 'CAPITAL_CALL', 'Additional capital call'],
        ['Global Private Equity Partners', '2023-03-15', 1800000, 'DISTRIBUTION', 'Partial distribution from portfolio company sale'],
        
        ['Prime Real Estate Fund IV', '2023-02-10', -600000, 'CAPITAL_CALL', 'Initial capital call'],
        ['Prime Real Estate Fund IV', '2023-06-30', -600000, 'CAPITAL_CALL', 'Second quarterly call'],
        
        ['Credit Opportunities Fund II', '2022-06-15', -1750000, 'CAPITAL_CALL', 'Initial funding for direct lending'],
        ['Credit Opportunities Fund II', '2022-12-20', -1750000, 'CAPITAL_CALL', 'Additional capital call'],
        ['Credit Opportunities Fund II', '2023-03-31', 175000, 'DISTRIBUTION', 'Quarterly interest distribution'],
        ['Credit Opportunities Fund II', '2023-06-30', 180000, 'DISTRIBUTION', 'Q2 interest distribution'],
        ['Credit Opportunities Fund II', '2023-09-30', 185000, 'DISTRIBUTION', 'Q3 interest distribution'],
        
        ['Emerging Markets Infrastructure', '2021-09-01', -2000000, 'CAPITAL_CALL', 'Initial capital commitment'],
        ['Emerging Markets Infrastructure', '2022-01-15', -1600000, 'CAPITAL_CALL', 'Deployment for renewable energy projects'],
        ['Emerging Markets Infrastructure', '2022-07-30', -1600000, 'CAPITAL_CALL', 'Transportation infrastructure investment'],
        ['Emerging Markets Infrastructure', '2023-04-15', -1200000, 'CAPITAL_CALL', 'Final deployment phase']
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row_idx, row_data in enumerate(cash_flows_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        adjusted_width = min(max_length + 2, 40)
        sheet.column_dimensions[column_name].width = adjusted_width
    
    # Save file
    output_path = 'sample_cashflows_data.xlsx'
    workbook.save(output_path)
    print(f"✅ Sample cash flows file created: {output_path}")
    return output_path

if __name__ == "__main__":
    print("Creating sample data files for testing...")
    holdings_file = create_sample_holdings()
    cashflows_file = create_sample_cash_flows()
    print(f"\n📁 Files created:")
    print(f"   Holdings: {holdings_file}")
    print(f"   Cash Flows: {cashflows_file}")
    print("\nYou can upload these files to quickly populate test data!")